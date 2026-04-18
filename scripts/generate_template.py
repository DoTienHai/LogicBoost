"""Generate Excel template for admin import."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


def generate_template():
    """Generate Excel template for question import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"
    
    columns = [
        "title", "title_vi",
        "question", "question_vi",
        "option_a", "option_b", "option_c", "option_d",
        "answer",
        "explanation", "explanation_vi",
        "mode", "sub_category", "difficulty", "time_limit"
    ]
    
    # Header styling
    header_fill = PatternFill(start_color="0277BD", end_color="0277BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Add headers
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Add a few example rows (empty but formatted)
    for row_idx in range(2, 5):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    
    # Set column widths
    column_widths = {
        "A": 20, "B": 18, "C": 30, "D": 28, "E": 12, "F": 12,
        "G": 12, "H": 12, "I": 12, "J": 35, "K": 32, "L": 16,
        "M": 14, "N": 12, "O": 12
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add data validation for dropdowns
    mode_dv = DataValidation(type="list", formula1='"daily_challenge,mini_game,real_world"', allow_blank=False)
    mode_dv.error = 'Select: daily_challenge, mini_game, or real_world'
    ws.add_data_validation(mode_dv)
    mode_dv.add("L2:L1000")
    
    sub_cat_dv = DataValidation(type="list", formula1='"finance,career,business"', allow_blank=True)
    sub_cat_dv.error = 'Select: finance, career, or business'
    ws.add_data_validation(sub_cat_dv)
    sub_cat_dv.add("M2:M1000")
    
    difficulty_dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    difficulty_dv.error = 'Select: 1 (Very Easy), 2 (Easy), 3 (Medium), 4 (Hard), or 5 (Very Hard)'
    ws.add_data_validation(difficulty_dv)
    difficulty_dv.add("N2:N1000")
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # ===== CREATE HELP SHEET =====
    help_ws = wb.create_sheet("Help")
    help_ws.append(["Column", "Type", "Valid Values", "Example"])
    
    help_data = [
        ["title", "Required", "English title (max 200 chars)", "Compound Interest Calculation"],
        ["title_vi", "Optional", "Vietnamese title (fallback to English)", "Tính Lãi Kép"],
        ["question", "Required", "English question (Markdown + LaTeX)", "You have **10M VNĐ** at **6%/year**. After 3 years: $$FV = PV × (1+r)^n$$"],
        ["question_vi", "Optional", "Vietnamese question (Markdown + LaTeX)", "Bạn có **10 triệu VNĐ**..."],
        ["option_a", "Optional", "Leave blank for free-text questions", "16,105,100"],
        ["option_b", "Optional", "Leave blank for free-text questions", "17,100,000"],
        ["option_c", "Optional", "Leave blank for free-text questions", "18,205,000"],
        ["option_d", "Optional", "Leave blank for free-text questions", "19,500,000"],
        ["answer", "Required", "a/b/c/d for MC, or exact text for free-text", "a"],
        ["explanation", "Required", "English explanation (Markdown + LaTeX)", "Using formula: $FV = 10M × (1.06)^3 = 11.91M$"],
        ["explanation_vi", "Optional", "Vietnamese explanation (Markdown + LaTeX)", "Sử dụng công thức..."],
        ["mode", "Required", "daily_challenge | mini_game | real_world", "daily_challenge"],
        ["sub_category", "Optional*", "finance | career | business (*required for real_world)", "finance"],
        ["difficulty", "Optional", "1 (🟢 Very Easy) | 2 (🟡 Easy) | 3 (🟠 Medium) | 4 (🔴 Hard) | 5 (⚫ Very Hard)", "2"],
        ["time_limit", "Optional", "Seconds (for mini_game mode only)", "60"]
    ]
    
    for row in help_data:
        help_ws.append(row)
    
    # Style help sheet header
    help_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    help_header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_idx in range(1, 5):
        cell = help_ws.cell(row=1, column=col_idx)
        cell.fill = help_header_fill
        cell.font = help_header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set help sheet column widths
    help_ws.column_dimensions['A'].width = 15
    help_ws.column_dimensions['B'].width = 12
    help_ws.column_dimensions['C'].width = 60
    help_ws.column_dimensions['D'].width = 40
    
    # Format help sheet rows
    for row_idx in range(2, len(help_data) + 2):
        for col_idx in range(1, 5):
            cell = help_ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    
    # Save file
    output_path = os.path.join(
        os.path.dirname(__file__),
        "..",
        "app", "static", "templates",
        "questions_template.xlsx"
    )
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Generated template: {output_path}")


if __name__ == "__main__":
    generate_template()
