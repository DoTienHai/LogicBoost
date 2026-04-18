"""Tests for import service."""
import pytest
import openpyxl
import os
import tempfile
from app.services.import_service import import_from_excel
from app.models import Question, db


@pytest.fixture
def sample_excel_file():
    """Create a sample Excel file for testing."""
    fd, filepath = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        
        headers = [
            "title", "title_vi",
            "question", "question_vi",
            "option_a", "option_b", "option_c", "option_d",
            "answer",
            "explanation", "explanation_vi",
            "mode", "sub_category", "difficulty", "time_limit"
        ]
        ws.append(headers)
        
        ws.append([
            "Question 1", "Câu hỏi 1",
            "What is 2+2?", "2 cộng 2 bằng mấy?",
            "4", "5", "6", "7",
            "a",
            "The answer is 4", "Câu trả lời là 4",
            "daily_challenge", "finance", "1", None
        ])
        
        ws.append([
            "Question 2", "Câu hỏi 2",
            "What is 3+3?", "3 cộng 3 bằng mấy?",
            "5", "6", "7", "8",
            "b",
            "The answer is 6", "Câu trả lời là 6",
            "mini_game", "business", "2", "60"
        ])
        
        wb.save(filepath)
        wb.close()
        
        yield filepath
    finally:
        import time
        time.sleep(0.1)
        try:
            if os.path.exists(filepath):
                os.remove(filepath)
        except:
            pass


class TestImportService:
    """Test import_from_excel function."""
    
    def test_import_valid_excel(self, app, sample_excel_file):
        """Test importing valid Excel file."""
        with app.app_context():
            results = import_from_excel(sample_excel_file)
            
            assert results['success'] == 2
            assert len(results['errors']) == 0
            
            q1 = Question.query.filter_by(title="Question 1").first()
            assert q1 is not None
            assert q1.mode == "daily_challenge"
            
            q2 = Question.query.filter_by(title="Question 2").first()
            assert q2 is not None
            assert q2.time_limit == 60
    
    def test_import_missing_title(self, app):
        """Test importing with missing title."""
        fd, filepath = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["title", "title_vi", "question", "question_vi", "option_a", "option_b", "option_c", "option_d", "answer", "explanation", "explanation_vi", "mode", "sub_category", "difficulty", "time_limit"]
            ws.append(headers)
            ws.append([None, None, "Q1", None, "a", "b", "c", "d", "a", "Exp", None, "daily_challenge", None, 1, None])
            wb.save(filepath)
            wb.close()
            
            with app.app_context():
                results = import_from_excel(filepath)
                assert results['success'] == 0
                assert len(results['errors']) > 0
                assert any("Title" in e for e in results['errors'])
        finally:
            import time
            time.sleep(0.1)
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except:
                pass
    
    def test_import_invalid_mode(self, app):
        """Test importing with invalid mode."""
        fd, filepath = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["title", "title_vi", "question", "question_vi", "option_a", "option_b", "option_c", "option_d", "answer", "explanation", "explanation_vi", "mode", "sub_category", "difficulty", "time_limit"]
            ws.append(headers)
            ws.append(["T1", None, "Q1", None, "a", "b", "c", "d", "a", "Exp", None, "invalid_mode", None, 1, None])
            wb.save(filepath)
            wb.close()
            
            with app.app_context():
                results = import_from_excel(filepath)
                assert results['success'] == 0
                assert len(results['errors']) > 0
                assert any("Invalid mode" in e for e in results['errors'])
        finally:
            import time
            time.sleep(0.1)
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except:
                pass
    
    def test_import_invalid_difficulty(self, app):
        """Test importing with invalid difficulty."""
        fd, filepath = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["title", "title_vi", "question", "question_vi", "option_a", "option_b", "option_c", "option_d", "answer", "explanation", "explanation_vi", "mode", "sub_category", "difficulty", "time_limit"]
            ws.append(headers)
            ws.append(["T1", None, "Q1", None, "a", "b", "c", "d", "a", "Exp", None, "daily_challenge", None, 6, None])
            wb.save(filepath)
            wb.close()
            
            with app.app_context():
                results = import_from_excel(filepath)
                assert results['success'] == 0
                assert len(results['errors']) > 0
                assert any("Difficulty" in e for e in results['errors'])
        finally:
            import time
            time.sleep(0.1)
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
            except:
                pass


class TestImportRoutes:
    """Test admin import routes."""
    
    def test_import_page_get(self, client):
        """Test import page GET."""
        response = client.get("/admin/import")
        assert response.status_code == 200
    
    def test_import_template(self, client):
        """Test template endpoint - should return Excel file."""
        response = client.get("/admin/import/template")
        assert response.status_code == 200
        # Check that response is Excel file
        assert "spreadsheetml" in response.content_type or "octet-stream" in response.content_type
    
    def test_import_no_file(self, client):
        """Test import without file."""
        response = client.post("/admin/import")
        assert response.status_code == 200
    
    def test_import_wrong_type(self, client):
        """Test import with wrong file type."""
        from io import BytesIO
        data = {'file': (BytesIO(b"test"), "test.txt")}
        response = client.post("/admin/import", data=data, content_type='multipart/form-data')
        assert response.status_code == 200
